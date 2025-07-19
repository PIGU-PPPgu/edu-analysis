#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
📊 XLSX转CSV转换器
使用标准库处理Excel文件，然后用现有的CSV分析脚本
"""

import sys
import csv
import json
from pathlib import Path

# 尝试导入openpyxl，如果失败则使用xml处理
try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    import zipfile
    import xml.etree.ElementTree as ET

def convert_xlsx_to_csv_openpyxl(xlsx_path, csv_path):
    """使用openpyxl转换XLSX到CSV"""
    workbook = load_workbook(xlsx_path, read_only=True)
    sheet = workbook.active
    
    with open(csv_path, 'w', newline='', encoding='utf-8') as csvfile:
        writer = csv.writer(csvfile)
        
        for row in sheet.iter_rows(values_only=True):
            # 将None值替换为空字符串
            cleaned_row = [str(cell) if cell is not None else '' for cell in row]
            writer.writerow(cleaned_row)
    
    workbook.close()
    print(f"✅ 成功转换: {xlsx_path} -> {csv_path}")

def convert_xlsx_to_csv_xml(xlsx_path, csv_path):
    """使用标准库XML处理转换XLSX到CSV"""
    print("⚠️ 使用简化XML处理器，可能无法处理复杂的Excel文件")
    
    # 基本的XML处理实现
    with zipfile.ZipFile(xlsx_path, 'r') as xlsx_zip:
        # 读取共享字符串
        shared_strings = []
        try:
            with xlsx_zip.open('xl/sharedStrings.xml') as f:
                root = ET.parse(f).getroot()
                for si in root.findall('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}si'):
                    t = si.find('{http://schemas.openxmlformats.org/spreadsheetml/2006/main}t')
                    if t is not None:
                        shared_strings.append(t.text or '')
                    else:
                        shared_strings.append('')
        except:
            print("⚠️ 无法读取共享字符串")
        
        # 读取工作表数据
        try:
            with xlsx_zip.open('xl/worksheets/sheet1.xml') as f:
                root = ET.parse(f).getroot()
                
                rows = []
                for row in root.findall('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}row'):
                    row_data = []
                    for cell in row.findall('{http://schemas.openxmlformats.org/spreadsheetml/2006/main}c'):
                        value = cell.find('{http://schemas.openxmlformats.org/spreadsheetml/2006/main}v')
                        if value is not None:
                            # 检查是否是共享字符串
                            t_attr = cell.get('t')
                            if t_attr == 's' and shared_strings:
                                try:
                                    idx = int(value.text)
                                    if 0 <= idx < len(shared_strings):
                                        row_data.append(shared_strings[idx])
                                    else:
                                        row_data.append(value.text or '')
                                except:
                                    row_data.append(value.text or '')
                            else:
                                row_data.append(value.text or '')
                        else:
                            row_data.append('')
                    rows.append(row_data)
                
                # 写入CSV
                with open(csv_path, 'w', newline='', encoding='utf-8') as csvfile:
                    writer = csv.writer(csvfile)
                    for row in rows:
                        writer.writerow(row)
                
                print(f"✅ 成功转换: {xlsx_path} -> {csv_path}")
                
        except Exception as e:
            print(f"❌ 转换失败: {e}")
            return False
    
    return True

def main():
    if len(sys.argv) != 2:
        print("❌ 使用方法: python xlsx-to-csv-converter.py <xlsx文件路径>")
        sys.exit(1)
    
    xlsx_path = Path(sys.argv[1])
    
    if not xlsx_path.exists():
        print(f"❌ 文件不存在: {xlsx_path}")
        sys.exit(1)
    
    if xlsx_path.suffix.lower() not in ['.xlsx', '.xls']:
        print(f"❌ 不支持的文件格式: {xlsx_path.suffix}")
        sys.exit(1)
    
    # 生成CSV文件路径
    csv_path = xlsx_path.with_suffix('.csv')
    
    print(f"📊 开始转换: {xlsx_path}")
    print(f"📁 目标文件: {csv_path}")
    
    # 选择转换方法
    if OPENPYXL_AVAILABLE:
        print("🔧 使用openpyxl转换器")
        convert_xlsx_to_csv_openpyxl(xlsx_path, csv_path)
    else:
        print("🔧 使用标准库XML转换器")
        convert_xlsx_to_csv_xml(xlsx_path, csv_path)
    
    print(f"✅ 转换完成: {csv_path}")
    print(f"💡 现在可以使用CSV分析脚本分析转换后的文件")

if __name__ == "__main__":
    main()